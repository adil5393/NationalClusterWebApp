"""Automated tests for redesigned executive XLSX reports and workbooks."""
import io
from datetime import datetime, timezone
import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app import models
from app.database import Base
from app.routers.reports import _build_full_workbook, _round_sheet_bytes
from app.routers.exports import export_participants_xlsx, export_room_allocation_xlsx


@pytest.fixture(scope="module")
def db_session():
    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)
    Session = sessionmaker(bind=engine)
    session = Session()

    # Seed test data
    t = models.Tournament(id=1, name="National Volleyball Championship 2026", age_group="U19 Boys")
    session.add(t)

    # Teams
    t1 = models.Team(id=1, name="Apex Strikers")
    t2 = models.Team(id=2, name="Coastal Titans")
    t3 = models.Team(id=3, name="Northern Knights")
    t4 = models.Team(id=4, name="Southern Eagles")
    session.add_all([t1, t2, t3, t4])

    # Venues
    v1 = models.Venue(id=1, name="Main Indoor Arena", location="Court A")
    session.add(v1)

    # Round 1: League Round with 2 Pools
    r1 = models.Round(id=1, tournament_id=1, name="Group Stage", sequence=1, format="LEAGUE")
    session.add(r1)

    p1 = models.Pool(id=1, tournament_id=1, round_id=1, name="Pool A", status="finalized")
    p1.teams = [t1, t2]
    session.add(p1)

    m1 = models.Match(
        id=1,
        tournament_id=1,
        round_id=1,
        match_type="LEAGUE",
        pool_id=1,
        team_a_id=1,
        team_b_id=2,
        team_a_score=3,
        team_b_score=1,
        winner_team_id=1,
        status="COMPLETED",
        venue_id=1,
        scheduled_at=datetime.now(timezone.utc),
    )
    session.add(m1)

    # Round 2: Knockout Finals
    r2 = models.Round(id=2, tournament_id=1, name="Championship Finals", sequence=2, format="KNOCKOUT")
    session.add(r2)

    m2 = models.Match(
        id=2,
        tournament_id=1,
        round_id=2,
        match_type="KNOCKOUT",
        team_a_id=1,
        team_b_id=3,
        team_a_score=0,
        team_b_score=0,
        status="SCHEDULED",
        venue_id=1,
        scheduled_at=datetime.now(timezone.utc),
    )
    m3 = models.Match(
        id=3,
        tournament_id=1,
        round_id=2,
        match_type="KNOCKOUT",
        team_a_id=4,
        team_b_id=None,
        status="COMPLETED",
        winner_team_id=4,
        notes="Bye",
    )
    session.add_all([m2, m3])

    # Participants
    p_athlete = models.Participant(id=1, team_id=1, full_name="John Doe", role="Player", gender="Male", age=17)
    p_coach = models.Participant(id=2, team_id=1, full_name="Coach Smith", role="Coach", gender="Male", age=45)
    session.add_all([p_athlete, p_coach])

    # Accommodation
    bld = models.Building(id=1, name="Block A", code="BLD-A")
    session.add(bld)
    flr = models.Floor(id=1, building_id=1, name="1st Floor", level=1)
    session.add(flr)
    rm = models.Room(id=1, floor_id=1, name="Room 101", capacity=4)
    session.add(rm)
    bed = models.Bed(id=1, room_id=1, label="Bed 1")
    session.add(bed)
    alloc = models.AccommodationAssignment(
        id=1,
        room_id=1,
        bed_id=1,
        participant_id=1,
        team_id=1,
    )
    session.add(alloc)

    session.commit()
    yield session
    session.close()


def test_round_snapshot_league_layout(db_session):
    t = db_session.get(models.Tournament, 1)
    r1 = db_session.get(models.Round, 1)
    
    file_bytes = _round_sheet_bytes(r1, t)
    assert file_bytes is not None
    assert len(file_bytes) > 0

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    assert "Group Stage" in wb.sheetnames
    ws = wb["Group Stage"]

    # Verify Banner
    assert "NATIONAL CLUSTER" in str(ws["A1"].value)
    assert "NATIONAL VOLLEYBALL CHAMPIONSHIP 2026" in str(ws["A2"].value)
    assert "OFFICIAL SNAPSHOT RECORD" in str(ws["A3"].value)

    # Verify KPI cards present
    assert ws["A5"].value == "League Stage"
    assert ws["A6"].value is not None

    # Search for Pool Standings & Fixtures
    values = [ws.cell(row=r, column=c).value for r in range(1, 35) for c in range(1, 13)]
    assert any("POOL STANDINGS & LEADERBOARD" in str(v) for v in values)
    assert any("Apex Strikers" in str(v) for v in values)
    assert any("M-1" in str(v) for v in values)


def test_round_snapshot_knockout_layout(db_session):
    t = db_session.get(models.Tournament, 1)
    r2 = db_session.get(models.Round, 2)
    
    file_bytes = _round_sheet_bytes(r2, t)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    assert "Championship Finals" in wb.sheetnames
    ws = wb["Championship Finals"]

    values = [ws.cell(row=r, column=c).value for r in range(1, 35) for c in range(1, 13)]
    assert any("KNOCKOUT BRACKET FIXTURES & RESULTS" in str(v) for v in values)
    assert any("M-2" in str(v) for v in values)
    assert any("M-3" in str(v) for v in values)
    assert any("BYE" in str(v) for v in values)


def test_full_tournament_workbook(db_session):
    t = db_session.get(models.Tournament, 1)
    file_bytes = _build_full_workbook(t)
    assert file_bytes is not None

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    assert "Tournament Overview" in wb.sheetnames
    assert "Group Stage" in wb.sheetnames
    assert "Championship Finals" in wb.sheetnames

    # Check Overview Sheet
    ws_ov = wb["Tournament Overview"]
    assert "TOURNAMENT MASTER OVERVIEW" in str(ws_ov["A3"].value)
    
    # Check Roadmap section
    ov_values = [ws_ov.cell(row=r, column=c).value for r in range(1, 30) for c in range(1, 11)]
    assert any("COMPETITION ROADMAP & ROUND STATUS" in str(v) for v in ov_values)
    assert any("Group Stage" in str(v) for v in ov_values)
    assert any("Championship Finals" in str(v) for v in ov_values)


def _read_stream(resp) -> bytes:
    import asyncio
    async def _collect():
        chunks = []
        async for chunk in resp.body_iterator:
            chunks.append(chunk)
        return b"".join(chunks)
    return asyncio.run(_collect())


def test_participants_and_rooms_xlsx_export(db_session):
    # Participants XLSX
    resp_p = export_participants_xlsx(db=db_session)
    p_bytes = _read_stream(resp_p)
    wb_p = openpyxl.load_workbook(io.BytesIO(p_bytes))
    assert "Participants Roster" in wb_p.sheetnames
    ws_p = wb_p["Participants Roster"]
    p_values = [ws_p.cell(row=r, column=c).value for r in range(1, 25) for c in range(1, 7)]
    assert any("John Doe" in str(v) for v in p_values)
    assert any("Coach Smith" in str(v) for v in p_values)

    # Room Allocation XLSX
    resp_r = export_room_allocation_xlsx(db=db_session)
    r_bytes = _read_stream(resp_r)
    wb_r = openpyxl.load_workbook(io.BytesIO(r_bytes))
    assert "Room Allocations" in wb_r.sheetnames
    ws_r = wb_r["Room Allocations"]
    r_values = [ws_r.cell(row=r, column=c).value for r in range(1, 25) for c in range(1, 7)]
    assert any("Block A" in str(v) for v in r_values)
    assert any("Room 101" in str(v) for v in r_values)
    assert any("John Doe" in str(v) for v in r_values)
