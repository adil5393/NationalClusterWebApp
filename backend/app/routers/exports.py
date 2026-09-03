"""Spreadsheet (CSV & Executive XLSX) exports for room allocation and participant lists."""
import csv
import io

import openpyxl
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from .. import models
from ..database import get_db
from ..excel_styler import (
    ALIGN_CENTER,
    ALIGN_HEADER_CENTER,
    ALIGN_HEADER_LEFT,
    ALIGN_LEFT,
    BORDER_CELL,
    BORDER_HEADER,
    FILL_TH_PRIMARY,
    FILL_ZEBRA_EVEN,
    FILL_ZEBRA_ODD,
    FONT_TD,
    FONT_TD_BOLD,
    FONT_TH,
    auto_fit_columns,
    enable_sheet_ergonomics,
    style_footer,
    style_header_banner,
    style_kpi_cards,
    style_section_bar,
)
from ..security import require_admin, require_module

router = APIRouter(prefix="/api/export", tags=["export"])

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _csv_response(header, rows, filename):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(header)
    writer.writerows(rows)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/participants.csv", dependencies=[Depends(require_module("teams"))])
def export_participants(db: Session = Depends(get_db)):
    teams = {t.id: t.name for t in db.query(models.Team).all()}
    rows = [
        [p.full_name, teams.get(p.team_id, ""), p.role or "", p.gender or "", p.age or ""]
        for p in db.query(models.Participant).order_by(models.Participant.team_id, models.Participant.full_name).all()
    ]
    return _csv_response(["Full Name", "Team", "Role", "Gender", "Age"], rows, "participants.csv")


@router.get("/participants.xlsx", dependencies=[Depends(require_module("teams"))])
def export_participants_xlsx(db: Session = Depends(get_db)):
    teams = {t.id: t.name for t in db.query(models.Team).all()}
    participants = db.query(models.Participant).order_by(models.Participant.team_id, models.Participant.full_name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participants Roster"
    max_cols = 6

    # Header Banner
    next_row = style_header_banner(
        ws,
        tournament_name="PARTICIPANTS & DELEGATE DIRECTORY",
        subtitle="Official Roster of Registered Athletes, Coaches & Staff",
        badge_text="OFFICIAL ROSTER EXPORT",
        max_col=max_cols,
        start_row=1,
    )

    # KPI Summary Cards
    total_p = len(participants)
    unique_teams_count = len({p.team_id for p in participants if p.team_id})
    players_count = sum(1 for p in participants if (p.role or "").lower() in ("player", "athlete", "student"))
    staff_count = total_p - players_count

    cards = [
        ("Total Participants", total_p, "Registered"),
        ("Teams Represented", unique_teams_count, "Affiliated Clubs"),
        ("Athletes / Players", players_count if players_count > 0 else total_p, "Competitors"),
        ("Staff / Coaches", staff_count if players_count > 0 else "—", "Officials"),
    ]
    next_row = style_kpi_cards(ws, cards, start_row=next_row, card_width_cols=1)

    # Table Section
    next_row = style_section_bar(ws, "Master Participant List", next_row, max_col=max_cols, icon="👥")

    headers = [
        ("NO.", 6, ALIGN_HEADER_CENTER),
        ("FULL NAME", 26, ALIGN_HEADER_LEFT),
        ("TEAM AFFILIATION", 24, ALIGN_HEADER_LEFT),
        ("ROLE", 14, ALIGN_HEADER_CENTER),
        ("GENDER", 10, ALIGN_HEADER_CENTER),
        ("AGE", 8, ALIGN_HEADER_CENTER),
    ]

    ws.row_dimensions[next_row].height = 22
    for col_idx, (th_label, _, align) in enumerate(headers, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=th_label)
        cell.font = FONT_TH
        cell.fill = FILL_TH_PRIMARY
        cell.alignment = align
        cell.border = BORDER_HEADER
    next_row += 1

    for idx, p in enumerate(participants, start=1):
        ws.row_dimensions[next_row].height = 20
        fill = FILL_ZEBRA_EVEN if idx % 2 == 0 else FILL_ZEBRA_ODD

        row_data = [
            (idx, ALIGN_CENTER, FONT_TD_BOLD),
            (p.full_name, ALIGN_LEFT, FONT_TD_BOLD),
            (teams.get(p.team_id, "—"), ALIGN_LEFT, FONT_TD),
            (p.role or "Player", ALIGN_CENTER, FONT_TD),
            (p.gender or "—", ALIGN_CENTER, FONT_TD),
            (p.age or "—", ALIGN_CENTER, FONT_TD),
        ]

        for col_idx, (val, align, font) in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.font = font
            cell.alignment = align
            cell.fill = fill
            cell.border = BORDER_CELL
        next_row += 1

    ws.row_dimensions[next_row].height = 12
    next_row += 1
    style_footer(ws, next_row, max_col=max_cols)

    auto_fit_columns(ws, min_width=8, max_width=45, extra_padding=3)
    enable_sheet_ergonomics(ws, freeze_pane="A7")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="participants_roster.xlsx"'},
    )


@router.get("/rooms.csv", dependencies=[Depends(require_module("accommodation"))])
def export_room_allocation(db: Session = Depends(get_db)):
    rows = []
    for a in db.query(models.AccommodationAssignment).all():
        room = a.room
        floor = room.floor if room else None
        building = floor.building if floor else None
        participant = db.get(models.Participant, a.participant_id) if a.participant_id else None
        rows.append([
            building.name if building else "",
            floor.name if floor else "",
            room.name if room else "",
            a.bed.label if a.bed else "",
            participant.full_name if participant else "(whole team)",
            a.team.name if a.team else "",
        ])
    return _csv_response(["Building", "Floor", "Room", "Bed", "Occupant", "Team"], rows, "room-allocation.csv")


@router.get("/rooms.xlsx", dependencies=[Depends(require_module("accommodation"))])
def export_room_allocation_xlsx(db: Session = Depends(get_db)):
    assignments = db.query(models.AccommodationAssignment).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Room Allocations"
    max_cols = 6

    next_row = style_header_banner(
        ws,
        tournament_name="ACCOMMODATION & ROOM ALLOCATION",
        subtitle="Building, Floor, Room, Bed & Assigned Occupant Details",
        badge_text="OFFICIAL ALLOCATION EXPORT",
        max_col=max_cols,
        start_row=1,
    )

    total_assignments = len(assignments)
    unique_buildings = len({a.room.floor.building_id for a in assignments if a.room and a.room.floor and a.room.floor.building_id})
    unique_rooms = len({a.room_id for a in assignments if a.room_id})
    individual_beds = sum(1 for a in assignments if a.bed_id is not None)

    cards = [
        ("Total Allocations", total_assignments, "Active Stays"),
        ("Buildings Utilized", unique_buildings, "Hostel Blocks"),
        ("Rooms Assigned", unique_rooms, "Occupied Rooms"),
        ("Bed Assignments", individual_beds, "Specific Beds"),
    ]
    next_row = style_kpi_cards(ws, cards, start_row=next_row, card_width_cols=1)

    next_row = style_section_bar(ws, "Master Room Allocation Schedule", next_row, max_col=max_cols, icon="🛏️")

    headers = [
        ("BUILDING", 20, ALIGN_HEADER_LEFT),
        ("FLOOR", 16, ALIGN_HEADER_LEFT),
        ("ROOM", 14, ALIGN_HEADER_CENTER),
        ("BED LABEL", 14, ALIGN_HEADER_CENTER),
        ("OCCUPANT NAME", 24, ALIGN_HEADER_LEFT),
        ("TEAM AFFILIATION", 24, ALIGN_HEADER_LEFT),
    ]

    ws.row_dimensions[next_row].height = 22
    for col_idx, (th_label, _, align) in enumerate(headers, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=th_label)
        cell.font = FONT_TH
        cell.fill = FILL_TH_PRIMARY
        cell.alignment = align
        cell.border = BORDER_HEADER
    next_row += 1

    for idx, a in enumerate(assignments, start=1):
        ws.row_dimensions[next_row].height = 20
        room = a.room
        floor = room.floor if room else None
        building = floor.building if floor else None
        participant = db.get(models.Participant, a.participant_id) if a.participant_id else None

        fill = FILL_ZEBRA_EVEN if idx % 2 == 0 else FILL_ZEBRA_ODD

        row_data = [
            (building.name if building else "—", ALIGN_LEFT, FONT_TD_BOLD),
            (floor.name if floor else "—", ALIGN_LEFT, FONT_TD),
            (room.name if room else "—", ALIGN_CENTER, FONT_TD_BOLD),
            (a.bed.label if a.bed else "(Any Bed)", ALIGN_CENTER, FONT_TD),
            (participant.full_name if participant else "(Whole Team)", ALIGN_LEFT, FONT_TD_BOLD),
            (a.team.name if a.team else "—", ALIGN_LEFT, FONT_TD),
        ]

        for col_idx, (val, align, font) in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.font = font
            cell.alignment = align
            cell.fill = fill
            cell.border = BORDER_CELL
        next_row += 1

    ws.row_dimensions[next_row].height = 12
    next_row += 1
    style_footer(ws, next_row, max_col=max_cols)

    auto_fit_columns(ws, min_width=8, max_width=45, extra_padding=3)
    enable_sheet_ergonomics(ws, freeze_pane="A7")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="room_allocations.xlsx"'},
    )


@router.get("/attendance.xlsx", dependencies=[Depends(require_module("attendance"))])
def export_attendance_xlsx(db: Session = Depends(get_db)):
    teams = {t.id: t.name for t in db.query(models.Team).all()}
    participants = (
        db.query(models.Participant)
        .order_by(models.Participant.age_group, models.Participant.team_id, models.Participant.full_name)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"
    max_cols = 6

    next_row = style_header_banner(
        ws,
        tournament_name="SQUAD ATTENDANCE REPORT",
        subtitle="Check-in Status for Every Registered Participant",
        badge_text="OFFICIAL ATTENDANCE EXPORT",
        max_col=max_cols,
        start_row=1,
    )

    total_p = len(participants)
    present_p = sum(1 for p in participants if p.is_present)
    absent_p = total_p - present_p
    rate = f"{round(100 * present_p / total_p)}%" if total_p else "—"

    cards = [
        ("Total Participants", total_p, "Registered"),
        ("Present", present_p, "Checked In"),
        ("Absent", absent_p, "Not Checked In"),
        ("Attendance Rate", rate, "Present / Total"),
    ]
    next_row = style_kpi_cards(ws, cards, start_row=next_row, card_width_cols=1)

    next_row = style_section_bar(ws, "Participant Attendance", next_row, max_col=max_cols, icon="✅")

    headers = [
        ("TEAM", 24, ALIGN_HEADER_LEFT),
        ("AGE GROUP", 14, ALIGN_HEADER_CENTER),
        ("PARTICIPANT NAME", 24, ALIGN_HEADER_LEFT),
        ("REG. NO.", 14, ALIGN_HEADER_CENTER),
        ("ROLE", 14, ALIGN_HEADER_CENTER),
        ("PRESENT", 12, ALIGN_HEADER_CENTER),
    ]

    ws.row_dimensions[next_row].height = 22
    for col_idx, (th_label, _, align) in enumerate(headers, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=th_label)
        cell.font = FONT_TH
        cell.fill = FILL_TH_PRIMARY
        cell.alignment = align
        cell.border = BORDER_HEADER
    next_row += 1

    for idx, p in enumerate(participants, start=1):
        ws.row_dimensions[next_row].height = 20
        fill = FILL_ZEBRA_EVEN if idx % 2 == 0 else FILL_ZEBRA_ODD

        row_data = [
            (teams.get(p.team_id, "—"), ALIGN_LEFT, FONT_TD_BOLD),
            (p.age_group or "—", ALIGN_CENTER, FONT_TD),
            (p.full_name, ALIGN_LEFT, FONT_TD_BOLD),
            (p.registration_no or "—", ALIGN_CENTER, FONT_TD),
            (p.role or "Player", ALIGN_CENTER, FONT_TD),
            ("PRESENT" if p.is_present else "ABSENT", ALIGN_CENTER, FONT_TD_BOLD),
        ]

        for col_idx, (val, align, font) in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.font = font
            cell.alignment = align
            cell.fill = fill
            cell.border = BORDER_CELL
        next_row += 1

    ws.row_dimensions[next_row].height = 12
    next_row += 1
    style_footer(ws, next_row, max_col=max_cols)

    auto_fit_columns(ws, min_width=8, max_width=45, extra_padding=3)
    enable_sheet_ergonomics(ws, freeze_pane="A7")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="attendance_report.xlsx"'},
    )


@router.get("/arrival.xlsx", dependencies=[Depends(require_module("teams"))])
def export_arrival_xlsx(db: Session = Depends(get_db)):
    teams = db.query(models.Team).order_by(models.Team.name).all()
    participant_counts = dict(
        db.query(models.Participant.team_id, func.count(models.Participant.id))
        .group_by(models.Participant.team_id)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Arrival Status"
    max_cols = 5

    next_row = style_header_banner(
        ws,
        tournament_name="TEAM ARRIVAL REPORT",
        subtitle="Delegation Arrival Status for Every Registered School",
        badge_text="OFFICIAL ARRIVAL EXPORT",
        max_col=max_cols,
        start_row=1,
    )

    total_t = len(teams)
    arrived_t = sum(1 for t in teams if t.has_arrived)
    not_arrived_t = total_t - arrived_t
    rate = f"{round(100 * arrived_t / total_t)}%" if total_t else "—"

    cards = [
        ("Total Teams", total_t, "Registered Schools"),
        ("Arrived", arrived_t, "Checked In"),
        ("Not Arrived", not_arrived_t, "Pending"),
        ("Arrival Rate", rate, "Arrived / Total"),
    ]
    next_row = style_kpi_cards(ws, cards, start_row=next_row, card_width_cols=1)

    next_row = style_section_bar(ws, "Delegation Arrival Status", next_row, max_col=max_cols, icon="🚌")

    headers = [
        ("SCHOOL / TEAM", 30, ALIGN_HEADER_LEFT),
        ("REGION", 18, ALIGN_HEADER_LEFT),
        ("COUNTRY", 14, ALIGN_HEADER_CENTER),
        ("ATHLETES", 12, ALIGN_HEADER_CENTER),
        ("ARRIVED", 12, ALIGN_HEADER_CENTER),
    ]

    ws.row_dimensions[next_row].height = 22
    for col_idx, (th_label, _, align) in enumerate(headers, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=th_label)
        cell.font = FONT_TH
        cell.fill = FILL_TH_PRIMARY
        cell.alignment = align
        cell.border = BORDER_HEADER
    next_row += 1

    for idx, t in enumerate(teams, start=1):
        ws.row_dimensions[next_row].height = 20
        fill = FILL_ZEBRA_EVEN if idx % 2 == 0 else FILL_ZEBRA_ODD

        row_data = [
            (t.name, ALIGN_LEFT, FONT_TD_BOLD),
            (t.region or "—", ALIGN_LEFT, FONT_TD),
            (t.country or "—", ALIGN_CENTER, FONT_TD),
            (participant_counts.get(t.id, 0), ALIGN_CENTER, FONT_TD),
            ("ARRIVED" if t.has_arrived else "NOT ARRIVED", ALIGN_CENTER, FONT_TD_BOLD),
        ]

        for col_idx, (val, align, font) in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.font = font
            cell.alignment = align
            cell.fill = fill
            cell.border = BORDER_CELL
        next_row += 1

    ws.row_dimensions[next_row].height = 12
    next_row += 1
    style_footer(ws, next_row, max_col=max_cols)

    auto_fit_columns(ws, min_width=8, max_width=45, extra_padding=3)
    enable_sheet_ergonomics(ws, freeze_pane="A7")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="arrival_report.xlsx"'},
    )


@router.get("/duties.xlsx", dependencies=[Depends(require_module("staff"))])
def export_duties_xlsx(db: Session = Depends(get_db)):
    duties = db.query(models.DutyAssignment).order_by(models.DutyAssignment.start_time.asc().nullslast()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Duty Roster"
    max_cols = 7

    next_row = style_header_banner(
        ws,
        tournament_name="STAFF DUTY REPORT",
        subtitle="Duty Assignments Across Every Building, Floor & Room",
        badge_text="OFFICIAL DUTY ROSTER EXPORT",
        max_col=max_cols,
        start_row=1,
    )

    total_d = len(duties)
    unique_staff = len({d.staff_id for d in duties})
    unique_duty_types = len({d.duty_type for d in duties if d.duty_type})
    unique_buildings = len({d.room.floor.building_id for d in duties if d.room and d.room.floor and d.room.floor.building_id})

    cards = [
        ("Total Assignments", total_d, "Duty Slots"),
        ("Staff Assigned", unique_staff, "Individuals"),
        ("Duty Types", unique_duty_types, "Categories"),
        ("Buildings Covered", unique_buildings, "Locations"),
    ]
    next_row = style_kpi_cards(ws, cards, start_row=next_row, card_width_cols=1)

    next_row = style_section_bar(ws, "Duty Roster", next_row, max_col=max_cols, icon="🛡️")

    headers = [
        ("STAFF NAME", 24, ALIGN_HEADER_LEFT),
        ("CATEGORY", 20, ALIGN_HEADER_LEFT),
        ("DUTY TYPE", 16, ALIGN_HEADER_CENTER),
        ("BUILDING", 18, ALIGN_HEADER_LEFT),
        ("ROOM", 14, ALIGN_HEADER_CENTER),
        ("START", 16, ALIGN_HEADER_CENTER),
        ("END", 16, ALIGN_HEADER_CENTER),
    ]

    ws.row_dimensions[next_row].height = 22
    for col_idx, (th_label, _, align) in enumerate(headers, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=th_label)
        cell.font = FONT_TH
        cell.fill = FILL_TH_PRIMARY
        cell.alignment = align
        cell.border = BORDER_HEADER
    next_row += 1

    for idx, d in enumerate(duties, start=1):
        ws.row_dimensions[next_row].height = 20
        fill = FILL_ZEBRA_EVEN if idx % 2 == 0 else FILL_ZEBRA_ODD
        room = d.room
        floor = room.floor if room else None
        building = floor.building if floor else None

        row_data = [
            (d.staff.full_name if d.staff else "—", ALIGN_LEFT, FONT_TD_BOLD),
            (d.staff.category if d.staff else "—", ALIGN_LEFT, FONT_TD),
            (d.duty_type or "—", ALIGN_CENTER, FONT_TD),
            (building.name if building else "—", ALIGN_LEFT, FONT_TD),
            (room.name if room else "—", ALIGN_CENTER, FONT_TD),
            (d.start_time.strftime("%d %b %Y %H:%M") if d.start_time else "—", ALIGN_CENTER, FONT_TD),
            (d.end_time.strftime("%d %b %Y %H:%M") if d.end_time else "—", ALIGN_CENTER, FONT_TD),
        ]

        for col_idx, (val, align, font) in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.font = font
            cell.alignment = align
            cell.fill = fill
            cell.border = BORDER_CELL
        next_row += 1

    ws.row_dimensions[next_row].height = 12
    next_row += 1
    style_footer(ws, next_row, max_col=max_cols)

    auto_fit_columns(ws, min_width=8, max_width=45, extra_padding=3)
    enable_sheet_ergonomics(ws, freeze_pane="A7")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="duty_report.xlsx"'},
    )


@router.get("/organizer-users.xlsx", dependencies=[Depends(require_admin)])
def export_organizer_users_xlsx(db: Session = Depends(get_db)):
    users = db.query(models.OrganizerUser).order_by(models.OrganizerUser.username).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Organizer Accounts"
    max_cols = 6

    next_row = style_header_banner(
        ws,
        tournament_name="ORGANIZER PORTAL USER REPORT",
        subtitle="Every Admin-Portal Login, Role & Module Permissions",
        badge_text="CONFIDENTIAL — ADMIN ONLY",
        max_col=max_cols,
        start_row=1,
    )

    total_u = len(users)
    admin_u = sum(1 for u in users if u.is_admin)
    active_u = sum(1 for u in users if u.is_active)
    inactive_u = total_u - active_u

    cards = [
        ("Total Accounts", total_u, "Organizer Logins"),
        ("Admins", admin_u, "Full Access"),
        ("Active", active_u, "Enabled"),
        ("Inactive", inactive_u, "Disabled"),
    ]
    next_row = style_kpi_cards(ws, cards, start_row=next_row, card_width_cols=1)

    next_row = style_section_bar(ws, "Organizer Portal Accounts", next_row, max_col=max_cols, icon="🔐")

    headers = [
        ("USERNAME", 20, ALIGN_HEADER_LEFT),
        ("FULL NAME", 22, ALIGN_HEADER_LEFT),
        ("ROLE", 14, ALIGN_HEADER_CENTER),
        ("STATUS", 12, ALIGN_HEADER_CENTER),
        ("MODULE PERMISSIONS", 40, ALIGN_HEADER_LEFT),
        ("LINKED STAFF", 24, ALIGN_HEADER_LEFT),
    ]

    ws.row_dimensions[next_row].height = 22
    for col_idx, (th_label, _, align) in enumerate(headers, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=th_label)
        cell.font = FONT_TH
        cell.fill = FILL_TH_PRIMARY
        cell.alignment = align
        cell.border = BORDER_HEADER
    next_row += 1

    for idx, u in enumerate(users, start=1):
        ws.row_dimensions[next_row].height = 20
        fill = FILL_ZEBRA_EVEN if idx % 2 == 0 else FILL_ZEBRA_ODD

        if u.is_admin:
            perms_display = "ALL MODULES (Admin)"
        else:
            perms_display = ", ".join(f"{k}: {v}" for k, v in (u.permissions or {}).items()) or "—"
        staff_display = ", ".join(s.full_name for s in u.staff_members) or "—"

        row_data = [
            (u.username, ALIGN_LEFT, FONT_TD_BOLD),
            (u.full_name or "—", ALIGN_LEFT, FONT_TD),
            ("Admin" if u.is_admin else "Staff", ALIGN_CENTER, FONT_TD_BOLD),
            ("ACTIVE" if u.is_active else "INACTIVE", ALIGN_CENTER, FONT_TD_BOLD),
            (perms_display, ALIGN_LEFT, FONT_TD),
            (staff_display, ALIGN_LEFT, FONT_TD),
        ]

        for col_idx, (val, align, font) in enumerate(row_data, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.font = font
            cell.alignment = align
            cell.fill = fill
            cell.border = BORDER_CELL
        next_row += 1

    ws.row_dimensions[next_row].height = 12
    next_row += 1
    style_footer(ws, next_row, max_col=max_cols)

    auto_fit_columns(ws, min_width=8, max_width=45, extra_padding=3)
    enable_sheet_ergonomics(ws, freeze_pane="A7")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": 'attachment; filename="organizer_users_report.xlsx"'},
    )
