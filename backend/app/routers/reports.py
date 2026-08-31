"""Match reports — executive-grade per-round snapshots and full-tournament master workbooks.

Features:
- Single-round snapshot reports with KPI metric cards, pool standings leaderboards, and styled match fixtures.
- Full tournament master workbooks with an executive Overview cover dashboard + detailed round tabs.
- High-contrast athletic theme with openpyxl styling, status pills, auto-fitting column widths, and freeze-panes.
"""
import io
from datetime import datetime, timezone
from typing import Dict, List, Set, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .. import models
from ..database import get_db
from ..excel_styler import (
    ALIGN_CENTER,
    ALIGN_HEADER_CENTER,
    ALIGN_HEADER_LEFT,
    ALIGN_HEADER_RIGHT,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    BORDER_CELL,
    BORDER_HEADER,
    BORDER_TOTAL_ROW,
    CLR_GRAY_TEXT,
    CLR_OBSIDIAN,
    CLR_SLATE_50,
    CLR_SLATE_600,
    CLR_WHITE,
    FILL_MEDAL_1,
    FILL_MEDAL_2,
    FILL_MEDAL_3,
    FILL_TH_LEAGUE,
    FILL_TH_PRIMARY,
    FILL_TH_SECONDARY,
    FILL_WINNER,
    FILL_ZEBRA_EVEN,
    FILL_ZEBRA_ODD,
    FONT_FAMILY,
    FONT_TD,
    FONT_TD_BOLD,
    FONT_TD_MUTED,
    FONT_TH,
    auto_fit_columns,
    enable_sheet_ergonomics,
    get_status_style,
    style_footer,
    style_header_banner,
    style_kpi_cards,
    style_section_bar,
)

router = APIRouter(prefix="/api", tags=["reports"])

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _round_format(round_: models.Round) -> str:
    if round_.format:
        return round_.format
    sample = round_.matches[0] if round_.matches else None
    return (sample.match_type if sample else None) or "KNOCKOUT"


def _safe_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no : \\ / ? * [ ]."""
    cleaned = "".join(c for c in (name or "") if c not in ':\\/?*[]').strip()
    return cleaned[:31] or "Round"


def _compute_pool_standings(pool: models.Pool) -> list[dict]:
    """Helper to compute pool standings if not already computed."""
    from .pools import compute_standings
    return compute_standings(pool)


def _render_round_sheet(
    ws: Worksheet,
    round_: models.Round,
    tournament: models.Tournament,
    is_snapshot: bool = True,
) -> None:
    """Builds a luxury executive layout for a single competition round."""
    fmt = _round_format(round_)
    is_league = fmt == "LEAGUE"
    matches = sorted(round_.matches, key=lambda m: ((m.pool.name if m.pool else ""), m.id))
    
    total_matches = len(matches)
    completed_matches = sum(1 for m in matches if m.status == "COMPLETED")
    live_matches = sum(1 for m in matches if m.status in ("LIVE", "IN_PROGRESS"))
    scheduled_matches = total_matches - completed_matches - live_matches

    max_cols = 12

    # 1. Top Banner
    badge = "OFFICIAL SNAPSHOT RECORD" if is_snapshot else "LIVE TOURNAMENT ROUND"
    sub_title = f"{round_.name}  |  Format: {'League / Round-Robin' if is_league else 'Knockout Elimination'}"
    if tournament.age_group:
        sub_title += f"  |  Category: {tournament.age_group}"
    
    next_row = style_header_banner(
        ws,
        tournament_name=tournament.name,
        subtitle=sub_title,
        badge_text=badge,
        max_col=max_cols,
        start_row=1,
    )

    # 2. KPI Cards (4 cards across 12 cols: 3 cols each)
    cards = [
        ("Competition Format", "League Stage" if is_league else "Knockout", "Official Rules"),
        ("Total Fixtures", total_matches, f"{completed_matches} Completed"),
        ("Completed", f"{completed_matches} / {total_matches}", f"{int(completed_matches / total_matches * 100) if total_matches > 0 else 0}% Done"),
        ("Live & Pending", f"{live_matches} Live • {scheduled_matches} Sched", "Upcoming"),
    ]
    next_row = style_kpi_cards(ws, cards, start_row=next_row, card_width_cols=3)

    # 3. If League: Render Pool Standings Tables
    if is_league and round_.pools:
        next_row = style_section_bar(ws, "Pool Standings & Leaderboard", next_row, max_col=max_cols, icon="🏆")
        
        # Standings Table Headers
        standings_headers = [
            ("POS", 6, ALIGN_HEADER_CENTER),
            ("POOL", 10, ALIGN_HEADER_CENTER),
            ("TEAM NAME", 26, ALIGN_HEADER_LEFT),
            ("PLAYED", 9, ALIGN_HEADER_CENTER),
            ("WON", 8, ALIGN_HEADER_CENTER),
            ("DRAWN", 8, ALIGN_HEADER_CENTER),
            ("LOST", 8, ALIGN_HEADER_CENTER),
            ("POINTS FOR", 12, ALIGN_HEADER_RIGHT),
            ("POINTS AGAINST", 14, ALIGN_HEADER_RIGHT),
            ("+/- DIFF", 10, ALIGN_HEADER_RIGHT),
            ("POINTS", 10, ALIGN_HEADER_CENTER),
            ("QUALIFICATION", 16, ALIGN_HEADER_CENTER),
        ]

        ws.row_dimensions[next_row].height = 22
        for col_idx, (th_label, _, align) in enumerate(standings_headers, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=th_label)
            cell.font = FONT_TH
            cell.fill = FILL_TH_LEAGUE
            cell.alignment = align
            cell.border = BORDER_HEADER
        next_row += 1

        # Render Standings Rows
        for pool in round_.pools:
            standings = _compute_pool_standings(pool)
            for row_idx, s in enumerate(standings):
                ws.row_dimensions[next_row].height = 20
                pos = s.get("position", row_idx + 1)
                diff = s.get("points_for", 0) - s.get("points_against", 0)
                diff_str = f"+{diff}" if diff > 0 else str(diff)

                # Qualification status hint
                qual_text = "Qualified (Rank 1)" if pos == 1 else ("Qualified (Rank 2)" if pos == 2 else "In Contention")

                row_vals = [
                    (pos, ALIGN_CENTER, FONT_TD_BOLD),
                    (pool.name, ALIGN_CENTER, FONT_TD),
                    (s.get("team_name", "Unknown"), ALIGN_LEFT, FONT_TD_BOLD),
                    (s.get("played", 0), ALIGN_CENTER, FONT_TD),
                    (s.get("won", 0), ALIGN_CENTER, FONT_TD),
                    (s.get("drawn", 0), ALIGN_CENTER, FONT_TD),
                    (s.get("lost", 0), ALIGN_CENTER, FONT_TD),
                    (s.get("points_for", 0), ALIGN_RIGHT, FONT_TD),
                    (s.get("points_against", 0), ALIGN_RIGHT, FONT_TD),
                    (diff_str, ALIGN_RIGHT, FONT_TD_BOLD),
                    (s.get("points", 0), ALIGN_CENTER, FONT_TD_BOLD),
                    (qual_text, ALIGN_CENTER, FONT_TD),
                ]

                # Medal highlight for top 3
                fill = FILL_ZEBRA_EVEN if row_idx % 2 == 0 else FILL_ZEBRA_ODD
                if pos == 1:
                    rank_fill = FILL_MEDAL_1
                elif pos == 2:
                    rank_fill = FILL_MEDAL_2
                elif pos == 3:
                    rank_fill = FILL_MEDAL_3
                else:
                    rank_fill = fill

                for col_idx, (val, align, font) in enumerate(row_vals, start=1):
                    cell = ws.cell(row=next_row, column=col_idx, value=val)
                    cell.font = font
                    cell.alignment = align
                    cell.fill = rank_fill if col_idx == 1 else fill
                    cell.border = BORDER_CELL
                next_row += 1

        # Spacing
        ws.row_dimensions[next_row].height = 12
        next_row += 1

    # 4. Match Fixtures & Live Results Table
    fixture_section_title = "Match Fixtures & Official Results" if is_league else "Knockout Bracket Fixtures & Results"
    next_row = style_section_bar(ws, fixture_section_title, next_row, max_col=max_cols, icon="⚔️")

    fixture_headers = [
        ("MATCH #", 9, ALIGN_HEADER_CENTER),
        ("STAGE / POOL", 14, ALIGN_HEADER_CENTER),
        ("TEAM A", 22, ALIGN_HEADER_LEFT),
        ("SCORE A", 9, ALIGN_HEADER_CENTER),
        ("VS", 5, ALIGN_HEADER_CENTER),
        ("SCORE B", 9, ALIGN_HEADER_CENTER),
        ("TEAM B", 22, ALIGN_HEADER_LEFT),
        ("WINNER / ADVANCED", 22, ALIGN_HEADER_LEFT),
        ("STATUS", 13, ALIGN_HEADER_CENTER),
        ("VENUE & COURT", 18, ALIGN_HEADER_LEFT),
        ("SCHEDULED TIME", 16, ALIGN_HEADER_CENTER),
        ("NOTES", 16, ALIGN_HEADER_LEFT),
    ]

    ws.row_dimensions[next_row].height = 22
    for col_idx, (th_label, _, align) in enumerate(fixture_headers, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=th_label)
        cell.font = FONT_TH
        cell.fill = FILL_TH_PRIMARY
        cell.alignment = align
        cell.border = BORDER_HEADER
    next_row += 1

    if not matches:
        ws.row_dimensions[next_row].height = 20
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=max_cols)
        empty_cell = ws.cell(row=next_row, column=1, value="No fixtures scheduled for this round yet.")
        empty_cell.font = FONT_TD_MUTED
        empty_cell.alignment = ALIGN_CENTER
        next_row += 1
    else:
        for m_idx, m in enumerate(matches):
            ws.row_dimensions[next_row].height = 20
            is_bye = m.notes == "Bye"
            stage_name = m.pool.name if m.pool else round_.name
            team_a_name = m.team_a.name if m.team_a else ("TBD" if not is_bye else "—")
            team_b_name = "— BYE —" if is_bye else (m.team_b.name if m.team_b else "TBD")
            
            score_a = m.team_a_score if (not is_bye and m.team_a_id and m.status == "COMPLETED") else ("—" if not is_bye else "")
            score_b = m.team_b_score if (not is_bye and m.team_b_id and m.status == "COMPLETED") else ("—" if not is_bye else "")
            
            winner_name = m.winner_team.name if m.winner_team else ("Team A (Bye)" if is_bye and m.team_a else ("—" if m.status != "COMPLETED" else "Tie / No Winner"))
            
            venue_info = f"{m.venue.name} ({m.venue.court_label})" if (m.venue and getattr(m.venue, "court_label", None)) else (m.venue.name if m.venue else "—")
            sched_time = m.scheduled_at.strftime("%d %b %H:%M") if m.scheduled_at else "—"
            status_text = "BYE" if is_bye else (m.status or "SCHEDULED")

            fill = FILL_ZEBRA_EVEN if m_idx % 2 == 0 else FILL_ZEBRA_ODD
            st_fill, st_font = get_status_style(status_text)

            row_data = [
                (f"M-{m.id}", ALIGN_CENTER, FONT_TD_BOLD, fill),
                (stage_name, ALIGN_CENTER, FONT_TD, fill),
                (team_a_name, ALIGN_LEFT, FONT_TD_BOLD, fill),
                (score_a, ALIGN_CENTER, FONT_TD_BOLD, fill),
                ("vs", ALIGN_CENTER, FONT_TD_MUTED, fill),
                (score_b, ALIGN_CENTER, FONT_TD_BOLD, fill),
                (team_b_name, ALIGN_LEFT, FONT_TD_BOLD, fill),
                (winner_name, ALIGN_LEFT, FONT_TD_BOLD if m.winner_team else FONT_TD, FILL_WINNER if m.winner_team else fill),
                (status_text, ALIGN_CENTER, st_font, st_fill),
                (venue_info, ALIGN_LEFT, FONT_TD, fill),
                (sched_time, ALIGN_CENTER, FONT_TD, fill),
                (m.notes or "—", ALIGN_LEFT, FONT_TD_MUTED, fill),
            ]

            for col_idx, (val, align, font, cell_fill) in enumerate(row_data, start=1):
                cell = ws.cell(row=next_row, column=col_idx, value=val)
                cell.font = font
                cell.alignment = align
                cell.fill = cell_fill
                cell.border = BORDER_CELL
            next_row += 1

    # Spacing and Footer
    ws.row_dimensions[next_row].height = 12
    next_row += 1
    style_footer(ws, next_row, max_col=max_cols)

    # Enable Ergonomics & Column Widths
    auto_fit_columns(ws, min_width=8, max_width=45, extra_padding=3)
    enable_sheet_ergonomics(ws, freeze_pane="A7")


def _render_overview_sheet(
    ws: Worksheet,
    tournament: models.Tournament,
) -> None:
    """Renders the executive master cover dashboard for full-tournament workbooks."""
    rounds = sorted(tournament.rounds, key=lambda r: r.sequence)
    max_cols = 10

    # 1. Header Banner
    sub = f"Comprehensive Multi-Round Competition Workbook  •  Age Group: {tournament.age_group or 'Open'}"
    next_row = style_header_banner(
        ws,
        tournament_name=tournament.name,
        subtitle=sub,
        badge_text="TOURNAMENT MASTER OVERVIEW",
        max_col=max_cols,
        start_row=1,
    )

    # 2. Aggregations
    total_rounds = len(rounds)
    all_matches = [m for r in rounds for m in r.matches]
    total_matches = len(all_matches)
    completed_matches = sum(1 for m in all_matches if m.status == "COMPLETED")
    live_matches = sum(1 for m in all_matches if m.status in ("LIVE", "IN_PROGRESS"))
    
    unique_team_ids = set()
    for m in all_matches:
        if m.team_a_id:
            unique_team_ids.add(m.team_a_id)
        if m.team_b_id:
            unique_team_ids.add(m.team_b_id)
    for r in rounds:
        for p in r.pools:
            for t in p.teams:
                unique_team_ids.add(t.id)

    total_teams_count = len(unique_team_ids)
    comp_pct = int(completed_matches / total_matches * 100) if total_matches > 0 else 0

    # KPI Summary Cards (across 10 cols)
    cards = [
        ("Competition Rounds", total_rounds, f"{total_matches} Total Fixtures"),
        ("Completed Fixtures", f"{completed_matches} / {total_matches}", f"{comp_pct}% Concluded"),
        ("Live & Active", live_matches, f"{total_matches - completed_matches - live_matches} Scheduled"),
        ("Competing Teams", total_teams_count, "Registered Teams"),
    ]
    next_row = style_kpi_cards(ws, cards, start_row=next_row, card_width_cols=2)

    # 3. Rounds Progression Roadmap Table
    next_row = style_section_bar(ws, "Competition Roadmap & Round Status", next_row, max_col=max_cols, icon="🗺️")
    
    roadmap_headers = [
        ("SEQ", 6, ALIGN_HEADER_CENTER),
        ("ROUND NAME", 26, ALIGN_HEADER_LEFT),
        ("FORMAT", 16, ALIGN_HEADER_CENTER),
        ("TOTAL MATCHES", 14, ALIGN_HEADER_CENTER),
        ("COMPLETED", 12, ALIGN_HEADER_CENTER),
        ("LIVE", 8, ALIGN_HEADER_CENTER),
        ("PENDING", 10, ALIGN_HEADER_CENTER),
        ("PROGRESS", 12, ALIGN_HEADER_CENTER),
        ("STATUS", 14, ALIGN_HEADER_CENTER),
        ("EXCEL TAB", 18, ALIGN_HEADER_LEFT),
    ]

    ws.row_dimensions[next_row].height = 22
    for col_idx, (th_label, _, align) in enumerate(roadmap_headers, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=th_label)
        cell.font = FONT_TH
        cell.fill = FILL_TH_PRIMARY
        cell.alignment = align
        cell.border = BORDER_HEADER
    next_row += 1

    if not rounds:
        ws.row_dimensions[next_row].height = 20
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=max_cols)
        empty_cell = ws.cell(row=next_row, column=1, value="No competition rounds created yet.")
        empty_cell.font = FONT_TD_MUTED
        empty_cell.alignment = ALIGN_CENTER
        next_row += 1
    else:
        for r_idx, r in enumerate(rounds):
            ws.row_dimensions[next_row].height = 20
            r_fmt = _round_format(r)
            r_matches = r.matches
            r_total = len(r_matches)
            r_comp = sum(1 for m in r_matches if m.status == "COMPLETED")
            r_live = sum(1 for m in r_matches if m.status in ("LIVE", "IN_PROGRESS"))
            r_pend = r_total - r_comp - r_live
            r_pct = f"{int(r_comp / r_total * 100)}%" if r_total > 0 else "0%"
            
            if r_comp == r_total and r_total > 0:
                r_status = "COMPLETED"
            elif r_live > 0 or r_comp > 0:
                r_status = "IN PROGRESS"
            else:
                r_status = "SCHEDULED"

            fill = FILL_ZEBRA_EVEN if r_idx % 2 == 0 else FILL_ZEBRA_ODD
            st_fill, st_font = get_status_style(r_status)
            tab_name = _safe_sheet_name(r.name)

            row_data = [
                (r.sequence, ALIGN_CENTER, FONT_TD_BOLD, fill),
                (r.name, ALIGN_LEFT, FONT_TD_BOLD, fill),
                ("League / Pool" if r_fmt == "LEAGUE" else "Knockout", ALIGN_CENTER, FONT_TD, fill),
                (r_total, ALIGN_CENTER, FONT_TD_BOLD, fill),
                (r_comp, ALIGN_CENTER, FONT_TD, fill),
                (r_live, ALIGN_CENTER, FONT_TD, fill),
                (r_pend, ALIGN_CENTER, FONT_TD, fill),
                (r_pct, ALIGN_CENTER, FONT_TD_BOLD, fill),
                (r_status, ALIGN_CENTER, st_font, st_fill),
                (f"Sheet: {tab_name}", ALIGN_LEFT, FONT_TD, fill),
            ]

            for col_idx, (val, align, font, cell_fill) in enumerate(row_data, start=1):
                cell = ws.cell(row=next_row, column=col_idx, value=val)
                cell.font = font
                cell.alignment = align
                cell.fill = cell_fill
                cell.border = BORDER_CELL
            next_row += 1

    # 4. Standings Summary across Pools (if tournament has any league rounds)
    all_pools = [p for r in rounds for p in r.pools]
    if all_pools:
        ws.row_dimensions[next_row].height = 12
        next_row += 1
        next_row = style_section_bar(ws, "Master Pool Standings & Leaderboard Summary", next_row, max_col=max_cols, icon="📊")

        pool_sum_headers = [
            ("RANK", 6, ALIGN_HEADER_CENTER),
            ("ROUND", 18, ALIGN_HEADER_LEFT),
            ("POOL", 12, ALIGN_HEADER_CENTER),
            ("TEAM NAME", 26, ALIGN_HEADER_LEFT),
            ("P", 7, ALIGN_HEADER_CENTER),
            ("W", 7, ALIGN_HEADER_CENTER),
            ("D", 7, ALIGN_HEADER_CENTER),
            ("L", 7, ALIGN_HEADER_CENTER),
            ("DIFF", 10, ALIGN_HEADER_RIGHT),
            ("TOTAL PTS", 12, ALIGN_HEADER_CENTER),
        ]

        ws.row_dimensions[next_row].height = 22
        for col_idx, (th_label, _, align) in enumerate(pool_sum_headers, start=1):
            cell = ws.cell(row=next_row, column=col_idx, value=th_label)
            cell.font = FONT_TH
            cell.fill = FILL_TH_LEAGUE
            cell.alignment = align
            cell.border = BORDER_HEADER
        next_row += 1

        row_counter = 0
        for p in all_pools:
            standings = _compute_pool_standings(p)
            for s in standings:
                ws.row_dimensions[next_row].height = 20
                pos = s.get("position", 1)
                diff = s.get("points_for", 0) - s.get("points_against", 0)
                diff_str = f"+{diff}" if diff > 0 else str(diff)

                fill = FILL_ZEBRA_EVEN if row_counter % 2 == 0 else FILL_ZEBRA_ODD
                rank_fill = FILL_MEDAL_1 if pos == 1 else (FILL_MEDAL_2 if pos == 2 else (FILL_MEDAL_3 if pos == 3 else fill))

                p_row_data = [
                    (pos, ALIGN_CENTER, FONT_TD_BOLD, rank_fill),
                    (p.round.name if p.round else "—", ALIGN_LEFT, FONT_TD, fill),
                    (p.name, ALIGN_CENTER, FONT_TD, fill),
                    (s.get("team_name", "Unknown"), ALIGN_LEFT, FONT_TD_BOLD, fill),
                    (s.get("played", 0), ALIGN_CENTER, FONT_TD, fill),
                    (s.get("won", 0), ALIGN_CENTER, FONT_TD, fill),
                    (s.get("drawn", 0), ALIGN_CENTER, FONT_TD, fill),
                    (s.get("lost", 0), ALIGN_CENTER, FONT_TD, fill),
                    (diff_str, ALIGN_RIGHT, FONT_TD_BOLD, fill),
                    (s.get("points", 0), ALIGN_CENTER, FONT_TD_BOLD, fill),
                ]

                for col_idx, (val, align, font, cell_fill) in enumerate(p_row_data, start=1):
                    cell = ws.cell(row=next_row, column=col_idx, value=val)
                    cell.font = font
                    cell.alignment = align
                    cell.fill = cell_fill
                    cell.border = BORDER_CELL
                next_row += 1
                row_counter += 1

    # Spacing and Footer
    ws.row_dimensions[next_row].height = 12
    next_row += 1
    style_footer(ws, next_row, max_col=max_cols)

    auto_fit_columns(ws, min_width=8, max_width=45, extra_padding=3)
    enable_sheet_ergonomics(ws, freeze_pane="A7")


def _round_sheet_bytes(round_: models.Round, tournament: models.Tournament) -> bytes:
    """Renders exactly one round into a beautifully styled one-sheet .xlsx snapshot."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(round_.name)
    _render_round_sheet(ws, round_, tournament, is_snapshot=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _build_full_workbook(tournament: models.Tournament) -> bytes:
    """Renders the entire tournament into an executive multi-sheet master workbook."""
    wb = openpyxl.Workbook()
    
    # Sheet 1: Overview
    overview_ws = wb.active
    overview_ws.title = "Tournament Overview"
    _render_overview_sheet(overview_ws, tournament)

    # Subsequent Sheets for each round
    rounds = sorted(tournament.rounds, key=lambda r: r.sequence)
    used_names: set[str] = {"Tournament Overview"}

    for r in rounds:
        base_name = _safe_sheet_name(r.name)
        sheet_name = base_name
        i = 2
        while sheet_name in used_names:
            suffix = f" ({i})"
            sheet_name = base_name[: 31 - len(suffix)] + suffix
            i += 1
        used_names.add(sheet_name)

        ws = wb.create_sheet(title=sheet_name)
        _render_round_sheet(ws, r, tournament, is_snapshot=False)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _report_dict(r: models.Report) -> dict:
    return {
        "id": r.id,
        "tournament_id": r.tournament_id,
        "round_id": r.round_id,
        "round_name": r.round_name,
        "round_sequence": r.round_sequence,
        "format": r.format,
        "generated_at": r.generated_at.isoformat(),
    }


@router.post("/tournaments/{tournament_id}/rounds/{round_id}/reports", status_code=201)
def generate_round_report(tournament_id: int, round_id: int, db: Session = Depends(get_db)):
    t = db.get(models.Tournament, tournament_id)
    if not t:
        raise HTTPException(404, "Tournament not found")
    r = db.get(models.Round, round_id)
    if not r or r.tournament_id != tournament_id:
        raise HTTPException(404, "Round not found for this tournament")

    file_bytes = _round_sheet_bytes(r, t)
    report = models.Report(
        tournament_id=tournament_id,
        round_id=round_id,
        round_name=r.name,
        round_sequence=r.sequence,
        format=_round_format(r),
        file_data=file_bytes,
        generated_at=datetime.now(timezone.utc),
    )
    db.add(report)
    db.commit()
    db.refresh(report)
    return _report_dict(report)


@router.get("/tournaments/{tournament_id}/reports")
def list_reports(tournament_id: int, db: Session = Depends(get_db)):
    if not db.get(models.Tournament, tournament_id):
        raise HTTPException(404, "Tournament not found")
    reports = (
        db.query(models.Report)
        .filter(models.Report.tournament_id == tournament_id)
        .order_by(models.Report.generated_at.desc())
        .all()
    )
    return [_report_dict(r) for r in reports]


@router.get("/reports/{report_id}/download")
def download_report(report_id: int, db: Session = Depends(get_db)):
    r = db.get(models.Report, report_id)
    if not r:
        raise HTTPException(404, "Report not found")
    filename = f"{_safe_sheet_name(r.round_name)}_report.xlsx"
    return StreamingResponse(
        iter([r.file_data]),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.delete("/reports/{report_id}", status_code=204)
def delete_report(report_id: int, db: Session = Depends(get_db)):
    r = db.get(models.Report, report_id)
    if not r:
        return
    db.delete(r)
    db.commit()


@router.get("/reports/tournaments/{tournament_id}/full.xlsx")
def download_full_report(tournament_id: int, db: Session = Depends(get_db)):
    t = db.get(models.Tournament, tournament_id)
    if not t:
        raise HTTPException(404, "Tournament not found")
    file_bytes = _build_full_workbook(t)
    filename = f"{_safe_sheet_name(t.name)}_full_report.xlsx"
    return StreamingResponse(
        iter([file_bytes]),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
