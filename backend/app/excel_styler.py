"""Excel styling engine and openpyxl formatting utilities for National Cluster reports.

Provides a unified corporate/athletic design system:
- High-contrast Obsidian & Gold executive palette
- KPI metric summary cards
- Visual section headers & branded top banners
- Styled data tables with zebra striping, status badges, medal rank highlights
- Auto-fitted column widths with padding, freeze panes, and visible gridlines
"""
from datetime import datetime, timezone
from typing import Any, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ----------------- DESIGN TOKENS ----------------- #
FONT_FAMILY = "Segoe UI"

# Colors
CLR_OBSIDIAN = "0F172A"       # Primary dark banner / table header
CLR_SLATE_800 = "1E293B"      # Secondary dark header
CLR_SLATE_700 = "334155"      # Muted dark
CLR_SLATE_600 = "475569"      # Subtitles
CLR_SLATE_400 = "94A3B8"      # Light borders / captions
CLR_SLATE_200 = "E2E8F0"      # Cell borders
CLR_SLATE_100 = "F1F5F9"      # Section header background / subtle accents
CLR_SLATE_50 = "F8FAFC"       # Zebra striping even rows
CLR_WHITE = "FFFFFF"          # White text / zebra odd rows

CLR_GOLD = "D97706"           # Brand accent gold (Amber 600)
CLR_GOLD_LIGHT = "FEF3C7"     # Gold highlight background (Amber 100)
CLR_GOLD_DARK = "92400E"      # Gold text (Amber 800)

CLR_ROYAL_BLUE = "1E3A8A"     # Blue 900 (Standings header)
CLR_BLUE_LIGHT = "E0F2FE"     # Sky 100
CLR_BLUE_DARK = "0369A1"      # Sky 700

CLR_GREEN_BG = "DCFCE7"       # Green 100 (Completed status)
CLR_GREEN_TEXT = "15803D"     # Green 700
CLR_AMBER_BG = "FEF3C7"       # Amber 100 (Live status)
CLR_AMBER_TEXT = "B45309"     # Amber 700
CLR_GRAY_BG = "F3F4F6"        # Gray 100 (Scheduled / Bye)
CLR_GRAY_TEXT = "6B7280"      # Gray 500
CLR_RED_BG = "FEE2E2"         # Red 100 (Cancelled)
CLR_RED_TEXT = "B91C1C"       # Red 700

# Fonts
FONT_BRAND = Font(name=FONT_FAMILY, size=9, bold=True, color="94A3B8")
FONT_TITLE = Font(name=FONT_FAMILY, size=15, bold=True, color="FFFFFF")
FONT_SUBTITLE = Font(name=FONT_FAMILY, size=10, bold=False, color="F8FAFC")
FONT_SECTION = Font(name=FONT_FAMILY, size=11, bold=True, color=CLR_OBSIDIAN)
FONT_TH = Font(name=FONT_FAMILY, size=10, bold=True, color=CLR_WHITE)
FONT_TH_SUB = Font(name=FONT_FAMILY, size=9, bold=True, color=CLR_WHITE)
FONT_TD = Font(name=FONT_FAMILY, size=9, color=CLR_OBSIDIAN)
FONT_TD_BOLD = Font(name=FONT_FAMILY, size=9, bold=True, color=CLR_OBSIDIAN)
FONT_TD_MUTED = Font(name=FONT_FAMILY, size=9, italic=True, color=CLR_GRAY_TEXT)
FONT_KPI_VAL = Font(name=FONT_FAMILY, size=16, bold=True, color=CLR_OBSIDIAN)
FONT_KPI_LBL = Font(name=FONT_FAMILY, size=8, bold=True, color=CLR_SLATE_600)
FONT_FOOTER = Font(name=FONT_FAMILY, size=8, italic=True, color="94A3B8")

# Fills
FILL_BANNER = PatternFill("solid", fgColor=CLR_OBSIDIAN)
FILL_BANNER_ACCENT = PatternFill("solid", fgColor=CLR_SLATE_800)
FILL_TH_PRIMARY = PatternFill("solid", fgColor=CLR_OBSIDIAN)
FILL_TH_LEAGUE = PatternFill("solid", fgColor=CLR_ROYAL_BLUE)
FILL_TH_SECONDARY = PatternFill("solid", fgColor=CLR_SLATE_800)
FILL_SECTION = PatternFill("solid", fgColor=CLR_SLATE_100)
FILL_ZEBRA_EVEN = PatternFill("solid", fgColor=CLR_SLATE_50)
FILL_ZEBRA_ODD = PatternFill("solid", fgColor=CLR_WHITE)
FILL_KPI_CARD = PatternFill("solid", fgColor=CLR_SLATE_50)
FILL_WINNER = PatternFill("solid", fgColor=CLR_GOLD_LIGHT)
FILL_MEDAL_1 = PatternFill("solid", fgColor="FEF08A")  # Gold 200
FILL_MEDAL_2 = PatternFill("solid", fgColor="E2E8F0")  # Slate 200
FILL_MEDAL_3 = PatternFill("solid", fgColor="FFEDD5")  # Orange 100

# Borders
BORDER_THIN_COLOR = CLR_SLATE_200
_side_thin = Side(style="thin", color=BORDER_THIN_COLOR)
_side_medium_obsidian = Side(style="medium", color=CLR_OBSIDIAN)
_side_gold = Side(style="medium", color=CLR_GOLD)
_side_double = Side(style="double", color=CLR_SLATE_700)

BORDER_CELL = Border(left=_side_thin, right=_side_thin, top=_side_thin, bottom=_side_thin)
BORDER_HEADER = Border(left=_side_thin, right=_side_thin, top=_side_thin, bottom=_side_medium_obsidian)
BORDER_KPI_CARD = Border(left=_side_thin, right=_side_thin, top=_side_gold, bottom=_side_thin)
BORDER_SECTION = Border(left=_side_gold, right=_side_thin, top=_side_thin, bottom=_side_thin)
BORDER_TOTAL_ROW = Border(left=_side_thin, right=_side_thin, top=_side_thin, bottom=_side_double)

# Alignments
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_HEADER_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_HEADER_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_HEADER_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def enable_sheet_ergonomics(ws: Worksheet, freeze_pane: str = "A7") -> None:
    """Configures freeze panes, visible gridlines, and default page setup."""
    try:
        if ws.views.sheetView:
            ws.views.sheetView[0].showGridLines = True
    except Exception:
        pass
    ws.freeze_panes = freeze_pane


def style_header_banner(
    ws: Worksheet,
    tournament_name: str,
    subtitle: str,
    badge_text: str = "OFFICIAL MATCH REPORT",
    max_col: int = 9,
    start_row: int = 1,
) -> int:
    """Renders a top executive banner across cols A through max_col.
    Returns the next available row number."""
    # Row 1: Brand watermark
    ws.row_dimensions[start_row].height = 18
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_col)
    top_cell = ws.cell(row=start_row, column=1, value="NATIONAL CLUSTER CHAMPIONSHIPS 2026-27")
    top_cell.font = FONT_BRAND
    top_cell.fill = FILL_BANNER
    top_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    for col in range(1, max_col + 1):
        ws.cell(row=start_row, column=col).fill = FILL_BANNER

    # Row 2: Main Tournament Title
    r2 = start_row + 1
    ws.row_dimensions[r2].height = 32
    ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=max_col)
    title_cell = ws.cell(row=r2, column=1, value=tournament_name.upper())
    title_cell.font = FONT_TITLE
    title_cell.fill = FILL_BANNER
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(1, max_col + 1):
        ws.cell(row=r2, column=col).fill = FILL_BANNER

    # Row 3: Subtitle / Round Info & Snapshot Type
    r3 = start_row + 2
    ws.row_dimensions[r3].height = 22
    ws.merge_cells(start_row=r3, start_column=1, end_row=r3, end_column=max_col)
    sub_text = f"{subtitle}  •  [{badge_text}]"
    sub_cell = ws.cell(row=r3, column=1, value=sub_text)
    sub_cell.font = FONT_SUBTITLE
    sub_cell.fill = FILL_BANNER_ACCENT
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in range(1, max_col + 1):
        ws.cell(row=r3, column=col).fill = FILL_BANNER_ACCENT

    # Row 4 is an empty spacing row
    ws.row_dimensions[start_row + 3].height = 10
    return start_row + 4


def style_kpi_cards(
    ws: Worksheet,
    cards: List[Tuple[str, Any, Optional[str]]],
    start_row: int,
    card_width_cols: int = 2,
) -> int:
    """Renders side-by-side metric cards (Label, Value, Optional Subtext).
    Returns the next available row number."""
    val_row = start_row
    lbl_row = start_row + 1
    ws.row_dimensions[val_row].height = 28
    ws.row_dimensions[lbl_row].height = 18

    curr_col = 1
    for label, val, sub in cards:
        end_col = curr_col + card_width_cols - 1
        ws.merge_cells(start_row=val_row, start_column=curr_col, end_row=val_row, end_column=end_col)
        v_cell = ws.cell(row=val_row, column=curr_col, value=val)
        v_cell.font = FONT_KPI_VAL
        v_cell.alignment = ALIGN_CENTER
        v_cell.fill = FILL_KPI_CARD

        ws.merge_cells(start_row=lbl_row, start_column=curr_col, end_row=lbl_row, end_column=end_col)
        lbl_display = f"{label.upper()} {f'({sub})' if sub else ''}"
        l_cell = ws.cell(row=lbl_row, column=curr_col, value=lbl_display)
        l_cell.font = FONT_KPI_LBL
        l_cell.alignment = ALIGN_CENTER
        l_cell.fill = FILL_KPI_CARD

        for c in range(curr_col, end_col + 1):
            ws.cell(row=val_row, column=c).fill = FILL_KPI_CARD
            ws.cell(row=lbl_row, column=c).fill = FILL_KPI_CARD
            ws.cell(row=val_row, column=c).border = Border(top=_side_gold, left=_side_thin if c == curr_col else None, right=_side_thin if c == end_col else None)
            ws.cell(row=lbl_row, column=c).border = Border(bottom=_side_thin, left=_side_thin if c == curr_col else None, right=_side_thin if c == end_col else None)

        curr_col = end_col + 1

    spacing_row = lbl_row + 1
    ws.row_dimensions[spacing_row].height = 12
    return spacing_row + 1


def style_section_bar(
    ws: Worksheet,
    title: str,
    row: int,
    max_col: int,
    icon: str = "▶",
) -> int:
    """Renders a section title divider with an accent line. Returns next row."""
    ws.row_dimensions[row].height = 24
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=f"{icon}  {title.upper()}")
    cell.font = FONT_SECTION
    cell.fill = FILL_SECTION
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.fill = FILL_SECTION
        c.border = Border(
            top=_side_thin,
            bottom=_side_thin,
            left=_side_gold if col == 1 else None,
            right=_side_thin if col == max_col else None,
        )

    return row + 1


def get_status_style(status_val: str) -> Tuple[PatternFill, Font]:
    """Returns background fill and font matching match/round status."""
    st = (status_val or "").strip().upper()
    if st in ("COMPLETED", "FINISHED", "FINAL"):
        return PatternFill("solid", fgColor=CLR_GREEN_BG), Font(name=FONT_FAMILY, size=9, bold=True, color=CLR_GREEN_TEXT)
    elif st in ("LIVE", "IN PROGRESS", "IN_PROGRESS"):
        return PatternFill("solid", fgColor=CLR_AMBER_BG), Font(name=FONT_FAMILY, size=9, bold=True, color=CLR_AMBER_TEXT)
    elif st in ("BYE", "WALKOVER"):
        return PatternFill("solid", fgColor=CLR_GRAY_BG), Font(name=FONT_FAMILY, size=9, italic=True, color=CLR_GRAY_TEXT)
    elif st in ("CANCELLED", "VOID"):
        return PatternFill("solid", fgColor=CLR_RED_BG), Font(name=FONT_FAMILY, size=9, bold=True, color=CLR_RED_TEXT)
    else:  # SCHEDULED / PENDING
        return PatternFill("solid", fgColor=CLR_GRAY_BG), Font(name=FONT_FAMILY, size=9, color=CLR_SLATE_600)


def auto_fit_columns(
    ws: Worksheet,
    min_width: int = 12,
    max_width: int = 42,
    extra_padding: int = 4,
    ignore_rows: Optional[set] = None,
) -> None:
    """Calculates optimal column width dynamically based on content."""
    if ignore_rows is None:
        ignore_rows = {1, 2, 3}  # Ignore merged banner rows

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.row in ignore_rows:
                continue
            if cell.coordinate in ws.merged_cells:
                continue
            if cell.value is not None:
                val_str = str(cell.value)
                lines = val_str.split("\n")
                line_max = max(len(l) for l in lines) if lines else 0
                max_len = max(max_len, line_max)

        width = max(min_width, min(max_len + extra_padding, max_width))
        ws.column_dimensions[col_letter].width = width


def style_footer(ws: Worksheet, row: int, max_col: int) -> int:
    """Writes a standardized executive footer with UTC timestamp."""
    ws.row_dimensions[row].height = 18
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    ts = datetime.now(timezone.utc).strftime("%d %b %Y %H:%M UTC")
    footer_text = f"Official Record generated by National Cluster System • {ts} • Confidential & Official"
    cell = ws.cell(row=row, column=1, value=footer_text)
    cell.font = FONT_FOOTER
    cell.alignment = ALIGN_CENTER
    return row + 1
